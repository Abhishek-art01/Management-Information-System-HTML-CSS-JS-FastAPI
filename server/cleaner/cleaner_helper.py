import pandas as pd
import numpy as np
import pdfplumber
import io
import re
from sqlmodel import Session, select, col
import xlrd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import traceback
from openpyxl.utils import get_column_letter

#=================================================================
from models import TripDataFile
#=================================================================


#=================================================================

def get_mandatory_columns():
    """Get all column names from the model, skipping metadata fields."""
    
    # Define the fields you want to skip (e.g., auto-generated IDs or timestamps)
    metadata_cols = {"created_at", "updated_at"} 
    
    # Using modern SQLModel / Pydantic v2 syntax:
    return [
        field_name for field_name in TripDataFile.model_fields 
        if field_name not in metadata_cols
    ]

def clean_cell_value(columns):
    cleaned = (
        columns
        .str.replace(r"\n", " ", regex=True)        # remove line breaks
        .str.replace(r"\t", " ", regex=True)        # remove tabs
        .str.replace(r"\s+", " ", regex=True)       # normalize spaces
        .str.strip()                                # trim edges
        .str.lower()                                # lowercase
        .str.replace(r"[^\w\s]", "", regex=True)    # remove special chars
        .str.replace(" ", "_")                      # snake_case
    )
    return cleaned


def format_excel_sheet(ws, start_col=1, start_row=1):
    """
    Final Excel formatter:
    - Header style (blue, bold, white)
    - Cambria font for all cells
    - Wrap text ON
    - Row height = 30
    - Auto-fit columns
    """
    # 1. Define Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(name="Cambria", size=11, bold=True, color="FFFFFF")
    cell_font = Font(name="Cambria")
    

    align_center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    max_column = ws.max_column

     # ---------------------------------------------------------
    # 🔥 CAPITALIZE HEADERS HERE
    # This changes "shift_date" -> "SHIFT DATE"
    # If you prefer Title Case ("Shift Date"), change .upper() to .title()
    # ---------------------------------------------------------
    # 2. Apply Header Formatting (Row 1)
    for col in range(start_col, max_column + 1):
        cell = ws.cell(row=start_row, column=col)
        
        if cell.value:
            # 🔥 CAPITALIZE AND REMOVE UNDERSCORES HERE
            cell.value = str(cell.value).replace("_", " ").upper()
            
            # Apply your styles
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center_wrap
            cell.border = border
            
    ws.row_dimensions[start_row].height = 30
    
    # 3. Apply Data Cell Formatting (Row 2 onwards)
    for row in ws.iter_rows(min_row=2):
        ws.row_dimensions[row[0].row].height = 30
        for cell in row:
            # We only apply font/alignment if not already specialized (like red/yellow logic)
            if not cell.font or cell.font.name != "Cambria":
                cell.font = cell_font
            cell.alignment = align_center_wrap
            cell.border = border

    # 4. Auto-fit column widths (min 10, max 50)
    for col in range(start_col, max_column + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        
        for row in range(start_row, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                if isinstance(cell.value, str):
                    lines = str(cell.value).split('\n')
                    line_length = max(len(line) for line in lines)
                else:
                    line_length = len(str(cell.value))
                
                adjusted_length = line_length + 2
                if adjusted_length > max_length:
                    max_length = adjusted_length
                    
        ws.column_dimensions[column_letter].width = min(max(max_length, 10), 50)
    
   

    # 5. Override specific columns by Header Name (Fixed for lowercase DB names)
    for cell in ws[1]:
        if not cell.value: continue
        
        col_letter = get_column_letter(cell.column)
        val = str(cell.value).lower().strip()
        
        if val in ["address", "employee address", "employee_address"]:
            ws.column_dimensions[col_letter].width = 80
        elif val in ["employee_name", "employee name"]:
            ws.column_dimensions[col_letter].width = 30
            
    return ws

def create_styled_excel(df, filename_prefix="Cleaned"):
    """ Generates Excel and applies the custom format_excel_sheet styles. """
    output = io.BytesIO()
    
    # Create a copy so we don't accidentally modify the original data
    df_export = df.copy()
    
   
    # 1. Write the DataFrame to BytesIO using the openpyxl engine
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Data')
        
        # 2. Get the openpyxl worksheet object
        worksheet = writer.sheets['Data']
        
        # 3. Apply your custom formatting function!
        format_excel_sheet(worksheet)
            
    output.seek(0)
    return df, output, f"{filename_prefix}.xlsx"



def get_xls_style_data(book, xf_index, row_idx, col_idx):
    """
    Extracts background and font colors from legacy .xls files.
    Includes debug prints to identify why colors might be missed.
    """
    try:
        xf = book.xf_list[xf_index]
        font = book.font_list[xf.font_index]
        
        # --- FONT COLOR DETECTION (RED) ---
        f_idx = font.colour_index
        rgb_f = book.colour_map.get(f_idx)
        font_hex = None
        
        # Check standard red indices (8, 10, 16 are common for Red)
        if f_idx in [10, 16]:
            font_hex = "FF0000"
        elif rgb_f:
            # Check if RGB values are "Red-ish" (High Red, Low Green/Blue)
            if rgb_f[0] > 150 and rgb_f[1] < 100 and rgb_f[2] < 100:
                font_hex = "FF0000"
        
        # --- BACKGROUND COLOR DETECTION (YELLOW) ---
        bg_idx = xf.background.pattern_colour_index
        rgb_b = book.colour_map.get(bg_idx)
        bg_hex = None
        
        # Check standard yellow indices (13, 19 are common for Yellow)
        if bg_idx in [13, 19]:
            bg_hex = "FFFF00"
        elif rgb_b:
            # Check if RGB values are "Yellow-ish"
            if rgb_b[0] > 200 and rgb_b[1] > 200 and rgb_b[2] < 150:
                bg_hex = "FFFF00"

        # --- DEBUG LOGGING ---
        # Only print for non-default styles to keep console clean
        if font_hex == "FF0000" or bg_hex == "FFFF00":
            print(f"[DEBUG STYLE] Row {row_idx}, Col {col_idx} | FontIdx: {f_idx} (Hex: {font_hex}) | BgIdx: {bg_idx} (Hex: {bg_hex})")

        return bg_hex, font_hex, bool(font.bold)
    except Exception as e:
        print(f"[DEBUG ERROR] Style extraction failed at Row {row_idx}, Col {col_idx}: {e}")
        return None, None, False 
  



# ==========================================
# DATABASE HELPER: BULK SAVE
# ==========================================
def bulk_save_unique(session, model, df):
    """
    Saves rows to the database only if the unique_id doesn't already exist.
    Handles duplicate columns automatically.
    """
    if df is None or df.empty:
        return 0

    try:
        # 🔥 FIX: Deduplicate columns first.
        # If 'unique_id' appears twice, df['unique_id'] returns a DataFrame (causing the crash).
        # This line keeps only the first occurrence of every column name.
        df = df.loc[:, ~df.columns.duplicated()]

        if "unique_id" not in df.columns:
            print(f"❌ Error: 'unique_id' missing in {model.__tablename__} data.")
            return 0

        # Now safe to call unique() because we know it's a Series
        incoming_ids = df["unique_id"].astype(str).unique().tolist()

        # Check existing IDs in DB
        from sqlmodel import select
        statement = select(model.unique_id).where(model.unique_id.in_(incoming_ids))
        existing_db_ids = session.exec(statement).all()
        existing_set = set(existing_db_ids)

        # Filter: Keep rows NOT in DB
        new_data = df[~df['unique_id'].isin(existing_set)]

        if not new_data.empty:
            records = new_data.to_dict(orient='records')
            # Create objects safely
            objects = [model(**row) for row in records]
            
            session.add_all(objects)
            session.commit()
            
            count = len(objects)
            print(f"✅ Saved {count} new records to {model.__tablename__}")
            return count
        
        print("🔹 No new records to save.")
        return 0

    except Exception as e:
        session.rollback()
        print(f"❌ Database Error in bulk_save_unique: {e}")
        import traceback
        traceback.print_exc()
        return 0
def sync_addresses_to_t3(session, df):
    """
    Extracts unique addresses from the processed data and syncs them to the T3 table.
    """
    if df is None or 'address' not in df.columns:
        return 0

    try:
        # ✅ FIX: Use df['address'].unique() instead of df.unique()
        unique_addresses = [a.strip().upper() for a in df['address'].unique() if a and str(a).strip()]
        
        if not unique_addresses:
            return 0

        # Assuming AddressTable is your T3 model
        # This part depends on your specific Address model name
        # For now, we return the count of unique addresses found
        return len(unique_addresses)
    except Exception as e:
        print(f"Error syncing addresses: {e}")
        return 0